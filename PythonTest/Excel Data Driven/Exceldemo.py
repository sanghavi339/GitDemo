import openpyxl

book = openpyxl.load_workbook("C:\\Users\\jay.s\\Desktop\\Pythondemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Jay Sanghavi"
print(sheet.cell(row=2, column=2).value)
#print(sheet.max_row) # To Find out Maximum Rows in excel sheet
#print(sheet.max_column) # to find out Maximum  Coulums in Excel sheet.

#print(sheet['A10'].value) # Print Perticular Row and Column Values.

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)